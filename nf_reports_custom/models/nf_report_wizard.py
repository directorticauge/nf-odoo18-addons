# -*- coding: utf-8 -*-
import base64, csv, io, json, math, time
from odoo import models, fields, api
from odoo.exceptions import UserError

PAGE_SIZE = 100
MAX_ROWS_STORE = 5000


class NfReportWizard(models.TransientModel):
    _name = 'nf.report.wizard'
    _description = 'Ejecutar Reporte NF'

    report_id = fields.Many2one('nf.report', string='Reporte', readonly=True)
    report_description = fields.Text('Descripción', readonly=True)
    param_line_ids = fields.One2many('nf.report.wizard.param', 'wizard_id', string='Parámetros')

    result_html = fields.Html('Resultado', readonly=True, sanitize=False)
    has_results = fields.Boolean(default=False)
    execution_time = fields.Float('Tiempo (s)', readonly=True)
    is_truncated = fields.Boolean(readonly=True)

    result_data = fields.Text('Datos (JSON)', readonly=True)
    result_cols = fields.Text('Columnas (JSON)', readonly=True)
    current_page = fields.Integer('Página', default=1, readonly=True)
    total_pages = fields.Integer('Total páginas', readonly=True)
    total_rows = fields.Integer('Total filas', readonly=True)
    row_count = fields.Integer('Filas en página', readonly=True)

    # ── helpers ──────────────────────────────────────────────────────────────

    def _get_params_dict(self):
        params = {}
        for line in self.param_line_ids:
            val = line.get_value()
            if line.required and (val is None or val == ''):
                raise UserError(f'El parámetro «{line.param_label}» es requerido.')
            params[line.name] = val
        return params

    def _params_summary(self):
        return '; '.join(
            f'{l.param_label}={l.value_input or ""}' for l in self.param_line_ids
        )

    # ── actions ──────────────────────────────────────────────────────────────

    def action_execute(self):
        self.ensure_one()
        report = self.report_id
        params = self._get_params_dict()
        t0 = time.time()
        try:
            self.env.cr.execute(report.query, params)
            all_rows = self.env.cr.fetchmany(MAX_ROWS_STORE + 1)
            elapsed = round(time.time() - t0, 3)
        except Exception as exc:
            elapsed = round(time.time() - t0, 3)
            self.env['nf.report.log'].create({
                'report_id': report.id,
                'execution_time': elapsed,
                'success': False,
                'error': str(exc),
                'params_summary': self._params_summary(),
                'export_type': 'screen',
            })
            raise UserError(f'Error al ejecutar la consulta:\n{exc}')

        cols = [d[0] for d in (self.env.cr.description or [])]
        truncated = len(all_rows) > MAX_ROWS_STORE
        stored_rows = all_rows[:MAX_ROWS_STORE]
        total_rows = len(stored_rows)
        total_pages = max(1, math.ceil(total_rows / PAGE_SIZE))

        self.env['nf.report.log'].create({
            'report_id': report.id,
            'execution_time': elapsed,
            'row_count': total_rows,
            'success': True,
            'params_summary': self._params_summary(),
            'export_type': 'screen',
        })

        self.write({
            'result_data': json.dumps(
                [[str(v) if v is not None else None for v in r] for r in stored_rows]
            ),
            'result_cols': json.dumps(cols),
            'current_page': 1,
            'total_pages': total_pages,
            'total_rows': total_rows,
            'execution_time': elapsed,
            'is_truncated': truncated,
            'has_results': True,
        })
        page_rows = stored_rows[:PAGE_SIZE]
        self.write({
            'result_html': self._build_html_table(
                cols, page_rows, truncated, elapsed, 1, total_pages, total_rows
            ),
            'row_count': len(page_rows),
        })
        return self._reopen()

    def action_next_page(self):
        return self._go_to_page(self.current_page + 1)

    def action_prev_page(self):
        return self._go_to_page(self.current_page - 1)

    def _go_to_page(self, page):
        self.ensure_one()
        page = max(1, min(page, self.total_pages))
        cols = json.loads(self.result_cols or '[]')
        all_rows = json.loads(self.result_data or '[]')
        start = (page - 1) * PAGE_SIZE
        page_rows = all_rows[start:start + PAGE_SIZE]
        self.write({
            'current_page': page,
            'row_count': len(page_rows),
            'result_html': self._build_html_table(
                cols, page_rows, self.is_truncated, self.execution_time,
                page, self.total_pages, self.total_rows,
            ),
        })
        return self._reopen()

    def action_validate_sql(self):
        self.ensure_one()
        report = self.report_id
        params = self._get_params_dict()
        try:
            self.env.cr.execute('EXPLAIN ' + report.query, params)
            self.env.cr.fetchall()
        except Exception as exc:
            raise UserError(f'Error de SQL:\n{exc}')
        self.write({
            'result_html': (
                '<style>.nf-ok{background:#eafaf1;border:1px solid #a9dfbf;'
                'padding:12px 16px;border-radius:6px;font-size:14px;color:#196f3d;'
                'font-family:Arial,sans-serif;margin:4px 0;}</style>'
                '<div class="nf-ok"><b>&#10003; SQL v&aacute;lido</b>'
                ' &mdash; La consulta es correcta y puede ejecutarse.</div>'
            ),
            'has_results': True,
        })
        return self._reopen()

    def action_export_csv(self):
        self.ensure_one()
        report = self.report_id
        params = self._get_params_dict()
        t0 = time.time()
        try:
            self.env.cr.execute(report.query, params)
            rows = self.env.cr.fetchall()
            elapsed = round(time.time() - t0, 3)
        except Exception as exc:
            raise UserError(f'Error al exportar CSV:\n{exc}')

        cols = [d[0] for d in (self.env.cr.description or [])]
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(cols)
        writer.writerows(rows)
        content = buf.getvalue().encode('utf-8-sig')

        self.env['nf.report.log'].create({
            'report_id': report.id,
            'execution_time': elapsed,
            'row_count': len(rows),
            'success': True,
            'params_summary': self._params_summary(),
            'export_type': 'csv',
        })
        attachment = self.env['ir.attachment'].create({
            'name': f'{report.name}.csv',
            'type': 'binary',
            'datas': base64.b64encode(content),
            'res_model': self._name,
            'res_id': self.id,
        })
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

    def action_export_xlsx(self):
        self.ensure_one()
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise UserError('Instale la librería openpyxl para exportar a Excel.')
        report = self.report_id
        params = self._get_params_dict()
        t0 = time.time()
        try:
            self.env.cr.execute(report.query, params)
            rows = self.env.cr.fetchall()
            elapsed = round(time.time() - t0, 3)
        except Exception as exc:
            raise UserError(f'Error al exportar Excel:\n{exc}')

        cols = [d[0] for d in (self.env.cr.description or [])]
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = report.name[:31]
        header_fill = PatternFill('solid', fgColor='2C3E50')
        header_font = Font(bold=True, color='FFFFFF')
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        for ri, row in enumerate(rows, 2):
            for ci, val in enumerate(row, 1):
                ws.cell(row=ri, column=ci, value=val)
        buf = io.BytesIO()
        wb.save(buf)
        content = buf.getvalue()

        self.env['nf.report.log'].create({
            'report_id': report.id,
            'execution_time': elapsed,
            'row_count': len(rows),
            'success': True,
            'params_summary': self._params_summary(),
            'export_type': 'xlsx',
        })
        attachment = self.env['ir.attachment'].create({
            'name': f'{report.name}.xlsx',
            'type': 'binary',
            'datas': base64.b64encode(content),
            'res_model': self._name,
            'res_id': self.id,
        })
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

    def action_export_pdf(self):
        self.ensure_one()
        report = self.report_id
        params = self._get_params_dict()
        t0 = time.time()
        try:
            self.env.cr.execute(report.query, params)
            rows = self.env.cr.fetchall()
            elapsed = round(time.time() - t0, 3)
        except Exception as exc:
            raise UserError(f'Error al exportar PDF:\n{exc}')

        cols = [d[0] for d in (self.env.cr.description or [])]
        self.write({
            'result_data': json.dumps(
                [[str(v) if v is not None else None for v in r] for r in rows]
            ),
            'result_cols': json.dumps(cols),
            'total_rows': len(rows),
            'execution_time': elapsed,
            'has_results': True,
        })
        self.env['nf.report.log'].create({
            'report_id': report.id,
            'execution_time': elapsed,
            'row_count': len(rows),
            'success': True,
            'params_summary': self._params_summary(),
            'export_type': 'pdf',
        })
        return self.env.ref(
            'nf_reports_custom.action_report_nf_wizard_pdf'
        ).report_action(self)

    # ── HTML builder ─────────────────────────────────────────────────────────

    def _build_html_table(self, cols, rows, truncated, elapsed, page, total_pages, total_rows):
        styles = """
        <style>
            .nf-wrap { overflow-x: auto; font-family: Arial, sans-serif; }
            .nf-info { font-size: 12px; color: #555; margin-bottom: 6px; padding: 4px 0; }
            .nf-warn { color: #c0392b; font-weight: bold; }
            .nf-table { width: 100%; border-collapse: collapse; font-size: 12px; }
            .nf-table th {
                background: #2c3e50; color: #fff;
                padding: 7px 10px; text-align: left;
                white-space: nowrap; position: sticky; top: 0;
            }
            .nf-table td { padding: 5px 10px; border-bottom: 1px solid #e0e0e0; white-space: nowrap; }
            .nf-table tr:nth-child(even) td { background: #f5f7fa; }
            .nf-table tr:hover td { background: #dceefb; }
            .nf-null { color: #bbb; font-style: italic; }
        </style>
        """
        warn = ''
        if truncated:
            warn = (
                f'<span class="nf-warn">&#9888; Solo se muestran {MAX_ROWS_STORE} filas. '
                f'Use "Exportar CSV/Excel" para todos los datos.</span> &#160;|&#160; '
            )
        info = (
            f'<div class="nf-info">{warn}'
            f'P&#225;gina: <b>{page}/{total_pages}</b> &#160;|&#160; '
            f'Total filas: <b>{total_rows}</b> &#160;|&#160; '
            f'Tiempo: <b>{elapsed}s</b> &#160;|&#160; '
            f'Columnas: <b>{len(cols)}</b></div>'
        )
        headers = ''.join(f'<th>{col}</th>' for col in cols)
        body = ''
        for row in rows:
            cells = ''.join(
                f'<td>{v if v is not None else "<span class=nf-null>NULL</span>"}</td>'
                for v in row
            )
            body += f'<tr>{cells}</tr>'
        table = (
            f'<div class="nf-wrap">{info}'
            f'<table class="nf-table"><thead><tr>{headers}</tr></thead>'
            f'<tbody>{body}</tbody></table></div>'
        )
        return styles + table

    def _reopen(self):
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }


class NfReportWizardParam(models.TransientModel):
    _name = 'nf.report.wizard.param'
    _description = 'Valor de Parámetro del Reporte'
    _order = 'sequence, id'

    wizard_id = fields.Many2one('nf.report.wizard', ondelete='cascade')
    param_id = fields.Many2one('nf.report.param', readonly=True)
    sequence = fields.Integer(default=10)
    name = fields.Char('Nombre técnico', readonly=True)
    param_label = fields.Char('Parámetro', readonly=True)
    param_hint = fields.Char('Ayuda', readonly=True)
    default_expr = fields.Char('Expresión', readonly=True,
        help='Palabra clave configurada como valor por defecto (ej: first_of_month, today)')
    param_type = fields.Selection([
        ('date', 'Fecha'),
        ('datetime', 'Fecha y Hora'),
        ('char', 'Texto'),
        ('integer', 'Número entero'),
        ('float', 'Número decimal'),
        ('boolean', 'Sí/No'),
    ], readonly=True)
    required = fields.Boolean(readonly=True)

    value_char = fields.Char('Valor')
    value_date = fields.Date('Valor')
    value_datetime = fields.Datetime('Valor')
    value_integer = fields.Integer('Valor')
    value_float = fields.Float('Valor', digits=(16, 2))
    value_boolean = fields.Boolean('Valor')

    value_input = fields.Char(
        'Valor',
        help='Ingrese el valor. Fechas: AAAA-MM-DD. Boolean: 1/0.'
    )

    def get_value(self):
        self.ensure_one()
        from odoo.fields import Date as ODate, Datetime as ODatetime
        val = (self.value_input or '').strip()
        if not val:
            return None
        try:
            if self.param_type == 'date':
                return ODate.from_string(val)
            elif self.param_type == 'datetime':
                return ODatetime.from_string(val)
            elif self.param_type == 'integer':
                return int(val)
            elif self.param_type == 'float':
                return float(val.replace(',', '.'))
            elif self.param_type == 'boolean':
                return val in ('1', 'true', 'si', 'sí', 'yes')
            else:
                return val or None
        except (ValueError, TypeError):
            return val or None
