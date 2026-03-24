#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script de importación de equipos en comodato desde Excel a Odoo (nf_support_cmd).

Hoja NEVERAS  → nf.cmd.equipment (tipo por nf.cmd.equipment.type)
Hoja DISPENSADORES → nf.cmd.equipment (tipo Dispensador)

Para cada equipo asignado se crea además un nf.cmd.assignment.

USO:
    pip install openpyxl
    python importar_comodatos.py
"""

import xmlrpc.client
import openpyxl
from datetime import datetime, date

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURACIÓN — ajusta estos valores
# ─────────────────────────────────────────────────────────────────────────────
ODOO_URL      = "https://indurod.augesoluciones.com"
ODOO_DB       = "indurod"          # nombre de la base de datos
ODOO_USER     = "admin"            # usuario administrador
ODOO_PASSWORD = "your_password"    # contraseña

EXCEL_PATH = r"C:\Users\NESTOR FORERO SALAS\Downloads\RELACION DE EQUIPOS - HIELO INDUROD S.A.S.   2025 -COPIA.xls"
# ─────────────────────────────────────────────────────────────────────────────

# Palabras clave que indican que el equipo está en patio (sin cliente)
DISPONIBLE_KEYWORDS = {
    "DISPONIBLE EN PATIO", "POR LOCALIZAR", "HINCAPIE JUAN CAMILO",
    "CARTAGENA",  # fila sin nombre real de cliente
}

def odoo_state_from_client(client_name: str) -> str:
    name = (client_name or "").strip().upper()
    if "POR LOCALIZAR" in name:
        return "lost"
    if "DISPONIBLE" in name:
        return "available"
    return "assigned"


def parse_date(raw) -> str | None:
    """Convierte varios formatos de fecha a 'YYYY-MM-DD' o None."""
    if not raw or str(raw).strip() in ("", "-"):
        return None
    raw = str(raw).strip()
    for fmt in ("%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(raw, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    # Excel can give serial numbers (already parsed by openpyxl as date)
    return None


def map_equipment_type_name(tipo: str) -> str:
    """Convierte el texto del Excel al nombre exacto en nf.cmd.equipment.type."""
    t = (tipo or "").strip().upper()
    if "CUARTO" in t:
        return "Cuarto Fr\u00edo"
    if "TERMO" in t:
        return "Termonevera"
    if "DISPENSADOR" in t or "WATER" in t:
        return "Dispensador"
    if "MEXIC" in t:
        return "Nevera Mexicana"
    if "CONGELADOR" in t:
        return "Congelador"
    if "NEVERA" in t or "STAR" in t:
        return "Nevera Star"
    return "Otro"


# Caché para no llamar a Odoo múltiples veces por el mismo tipo
_type_id_cache: dict = {}


def get_equipment_type_id(models_obj, uid, tipo_raw: str, explicit_name: str = None) -> int:
    """Busca el ID del tipo en Odoo (nf.cmd.equipment.type) y lo cachea."""
    name = explicit_name or map_equipment_type_name(tipo_raw)
    if name in _type_id_cache:
        return _type_id_cache[name]
    results = models_obj.execute_kw(
        ODOO_DB, uid, ODOO_PASSWORD,
        'nf.cmd.equipment.type', 'search_read',
        [[['name', '=', name]]],
        {'fields': ['id'], 'limit': 1},
    )
    if results:
        _type_id_cache[name] = results[0]['id']
        return _type_id_cache[name]
    # Fallback a "Otro" si el tipo no existe
    fallback = models_obj.execute_kw(
        ODOO_DB, uid, ODOO_PASSWORD,
        'nf.cmd.equipment.type', 'search_read',
        [[['name', '=', 'Otro']]],
        {'fields': ['id'], 'limit': 1},
    )
    fallback_id = fallback[0]['id'] if fallback else False
    _type_id_cache[name] = fallback_id
    return fallback_id


def connect_odoo():
    common = xmlrpc.client.ServerProxy(f"{ODOO_URL}/xmlrpc/2/common")
    uid = common.authenticate(ODOO_DB, ODOO_USER, ODOO_PASSWORD, {})
    if not uid:
        raise RuntimeError("❌ Autenticación fallida. Revisa usuario/contraseña.")
    models = xmlrpc.client.ServerProxy(f"{ODOO_URL}/xmlrpc/2/object")
    print(f"✅ Conectado a Odoo como uid={uid}")
    return uid, models


def create_equipment(models, uid, vals: dict) -> int:
    return models.execute_kw(
        ODOO_DB, uid, ODOO_PASSWORD,
        "nf.cmd.equipment", "create", [vals]
    )


def create_assignment(models, uid, vals: dict) -> int:
    return models.execute_kw(
        ODOO_DB, uid, ODOO_PASSWORD,
        "nf.cmd.assignment", "create", [vals]
    )


def import_neveras(ws, models, uid):
    print("\n📋 Importando hoja NEVERAS...")
    ok = 0
    skip = 0
    errors = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Columnas: 0=verif,1=ex_fis,2=N°,3=CLIENTE,4=NOMBRE COMERCIAL,
        #           5=NIT,6=FECHA,7=RUTA,8=CIUDAD,9=OK,10=CAPACIDAD,
        #           11=TIPO,12=LITROS,13=N/C
        try:
            num_equipo   = row[2]
            client_raw   = str(row[3] or "").strip()
            commercial   = str(row[4] or "").strip()
            nit          = str(row[5] or "").strip()
            fecha_raw    = row[6]
            ruta         = str(row[7] or "").strip()
            ciudad       = str(row[8] or "").strip()
            capacidad    = str(row[10] or "").strip()
            tipo_raw     = str(row[11] or "").strip()
            litros       = str(row[12] or "").strip()
            nc           = str(row[13] or "").strip()

            if not num_equipo or not tipo_raw:
                skip += 1
                continue

            equip_type_id = get_equipment_type_id(models, uid, tipo_raw)
            estado      = odoo_state_from_client(client_raw or commercial)
            fecha       = parse_date(fecha_raw) if isinstance(fecha_raw, str) else (
                fecha_raw.strftime("%Y-%m-%d") if hasattr(fecha_raw, "strftime") else None
            )

            equip_vals = {
                "sequence_number": int(num_equipo),
                "equipment_type_id": equip_type_id,
                "model_ref":       nc or None,
                "capacity":        capacidad or None,
                "liters":          litros or None,
                "state":           estado,
                "notes":           f"Importado desde Excel. Tipo original: {tipo_raw}",
            }
            equip_id = create_equipment(models, uid, equip_vals)

            # Crear asignación solo si hay un cliente real
            if estado == "assigned" and commercial and commercial.upper() not in DISPONIBLE_KEYWORDS:
                assign_vals = {
                    "equipment_id":    equip_id,
                    "commercial_name": commercial,
                    "nit_cc":          nit or None,
                    "route":           ruta or None,
                    "city":            ciudad or None,
                    "state":           "active",
                }
                if fecha:
                    assign_vals["delivery_date"] = fecha
                create_assignment(models, uid, assign_vals)

            ok += 1
            print(f"  R{row_idx}: N°{num_equipo} {map_equipment_type_name(tipo_raw)} — {commercial or 'DISPONIBLE'} ✓")

        except Exception as e:
            errors.append(f"  R{row_idx}: {e}")
            print(f"  R{row_idx}: ⚠️  {e}")

    print(f"\n  ✅ NEVERAS: {ok} equipos creados, {skip} filas vacías saltadas, {len(errors)} errores.")
    return errors


def import_dispensadores(ws, models, uid):
    print("\n📋 Importando hoja DISPENSADORES...")
    ok = 0
    skip = 0
    errors = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Columnas: 0=VERIF,1=N°,2=CLIENTE,3=NOMBRE COMERCIAL,4=NIT,
        #           5=RUTA,6=MARCA,7=FECHA,8=BOT,9=OK
        try:
            num_equipo  = row[1]
            client_raw  = str(row[2] or "").strip()
            commercial  = str(row[3] or "").strip()
            nit         = str(row[4] or "").strip()
            ruta        = str(row[5] or "").strip()
            marca       = str(row[6] or "").strip()
            fecha_raw   = row[7]
            bot         = str(row[8] or "").strip()

            if not num_equipo:
                skip += 1
                continue

            estado = odoo_state_from_client(client_raw or commercial)
            fecha = parse_date(fecha_raw) if isinstance(fecha_raw, str) else (
                fecha_raw.strftime("%Y-%m-%d") if hasattr(fecha_raw, "strftime") else None
            )

            notes_parts = []
            if marca:
                notes_parts.append(f"Marca/Serial: {marca}")
            if bot:
                notes_parts.append(f"Botellones: {bot}")
            notes_parts.append("Importado desde Excel.")

            equip_vals = {
                "sequence_number": int(num_equipo),
                "equipment_type_id": get_equipment_type_id(models, uid, "", explicit_name="Dispensador"),
                "brand":           marca or None,
                "state":           estado,
                "notes":           " | ".join(notes_parts),
            }
            equip_id = create_equipment(models, uid, equip_vals)

            if estado == "assigned" and commercial and commercial.upper() not in DISPONIBLE_KEYWORDS:
                assign_vals = {
                    "equipment_id":    equip_id,
                    "commercial_name": commercial,
                    "nit_cc":          nit or None,
                    "route":           ruta or None,
                    "state":           "active",
                }
                if fecha:
                    assign_vals["delivery_date"] = fecha
                create_assignment(models, uid, assign_vals)

            ok += 1
            print(f"  R{row_idx}: N°{num_equipo} Dispensador — {commercial or 'DISPONIBLE'} ✓")

        except Exception as e:
            errors.append(f"  R{row_idx}: {e}")
            print(f"  R{row_idx}: ⚠️  {e}")

    print(f"\n  ✅ DISPENSADORES: {ok} equipos creados, {skip} filas vacías saltadas, {len(errors)} errores.")
    return errors


def main():
    print("=" * 60)
    print("  IMPORTACIÓN DE COMODATOS — HIELO INDUROD")
    print("=" * 60)

    # Conectar a Odoo
    uid, models = connect_odoo()

    # Abrir Excel
    print(f"\n📂 Abriendo Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    print(f"   Hojas encontradas: {wb.sheetnames}")

    all_errors = []

    if "NEVERAS" in wb.sheetnames:
        errs = import_neveras(wb["NEVERAS"], models, uid)
        all_errors.extend(errs)
    else:
        print("⚠️  Hoja NEVERAS no encontrada.")

    if "DISPENSADORES" in wb.sheetnames:
        errs = import_dispensadores(wb["DISPENSADORES"], models, uid)
        all_errors.extend(errs)
    else:
        print("⚠️  Hoja DISPENSADORES no encontrada.")

    print("\n" + "=" * 60)
    if all_errors:
        print(f"  ⚠️  {len(all_errors)} errores durante la importación:")
        for e in all_errors:
            print(e)
    else:
        print("  🎉 Importación completada sin errores.")
    print("=" * 60)


if __name__ == "__main__":
    main()
